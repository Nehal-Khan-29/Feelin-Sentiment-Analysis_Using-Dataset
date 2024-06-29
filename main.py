import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import ttk as ttk
from PIL import ImageTk,Image
import pandas as pd
import openpyxl
from openpyxl import Workbook
import os
from textblob import TextBlob
import nltk
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report








nltk.download('stopwords')

try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download("punkt")

try:
    from textblob.download_corpora import main
    main()
except nltk.exceptions.ContentRetrievalError:
    print("Error downloading corpora. Please check your internet connection.")






# Page Close Confirmations (Messagebox):
def homeclose():
    if messagebox.askokcancel('Quit','Do you want to logout and quit?'):
        home.destroy()
        quit()




    
    
# icon window
icon = tk.Tk()
icon.title('Feelin')
icon.iconbitmap("logo_ico.ico")
image = Image.open("logo.png")
tk_image = ImageTk.PhotoImage(image)
image_label = tk.Label(icon, image=tk_image)
image_label.pack()
icon.update()
screen_width = icon.winfo_screenwidth()
screen_height = icon.winfo_screenheight()
window_width = 447  
window_height = 447  
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
icon.geometry("+{}+{}".format(x, y))
icon.after(2000, icon.destroy)
icon.mainloop()







#Excel add
filename = "sentiments.xlsx"

if not os.path.exists(filename):
    df = pd.DataFrame({'Positive':[],'Neutral':[],'Negative':[]})
    with pd.ExcelWriter(filename) as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    print(f"Excel file '{filename}' created.")
else:
    print(f"Excel file '{filename}' already exists.")
    
def resetpoints():
    global filename
    filename = "sentiments.xlsx"
    os.remove(filename)
    df = pd.DataFrame({'Positive':[],'Neutral':[],'Negative':[]})
    with pd.ExcelWriter(filename) as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    home2.destroy()

   
   
   
   
            
            
    
# Generate   
def generate():
    user_text = prompt_text_widget1.get("1.0", "end-1c")
    
    def analyze_sentiment(text):
        polarity = TextBlob(text).sentiment.polarity
        
        if polarity == 0.00:
            sentiment = 'Neutral'
            accuracy = 100
        else:  
            df = pd.read_csv('sentiment_Dataset.csv')
            stop_words = set(stopwords.words('english'))
            def preprocess_text(text):
                text = text.lower()
                text = ''.join([char for char in text if char.isalnum() or char.isspace()])
                words = text.split()
                words = [word for word in words if word not in stop_words]
                return ' '.join(words)
            df['New_sentence'] = df['review'].apply(preprocess_text)
            vectorizer = TfidfVectorizer()
            X = vectorizer.fit_transform(df['New_sentence'])
            y = df['sentiment']
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
            model = LogisticRegression()
            model.fit(X_train, y_train)
            y_pred = model.predict(X_test)
            accuracy = accuracy_score(y_test, y_pred)
            
            cleaned_input = preprocess_text(user_text)
            input_vector = vectorizer.transform([cleaned_input])
            sentiment = model.predict(input_vector)[0]

        return sentiment, polarity, accuracy 
        
    sentiment, polarity,accuracy = analyze_sentiment(user_text)
    if accuracy == 100:
        output = f"Sentiment: {sentiment}\nPolarity: {polarity:.2f}"
    else:
        output = f"Sentiment: {sentiment}\nAccuracy: {accuracy:.2f}"
    prompt_text_widget2.config(state=NORMAL)
    prompt_text_widget2.delete("1.0", END)
    prompt_text_widget2.insert(END,output)
    prompt_text_widget2.config(state=DISABLED)
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    col_num = 2
    
    if sentiment == "positive":
        col_num = 1
    elif sentiment == "negative":
        col_num = 3
    elif sentiment == "Neutral":
        col_num = 2
    
    row_num = 1

    while sheet.cell(row=row_num, column=col_num).value:
        row_num += 1
    sheet.cell(row=row_num, column=col_num).value = user_text
    wb.save(filename)

    
    
            
          
        
        
        
          
              
# Analysis
def analysis_window():
    global prompt_text_widget1,prompt_text_widget2
    
    prompt_page = tk.Tk()
    prompt_page.geometry('500x500')
    prompt_page.title('Feelin - Analysis')
    prompt_page.iconbitmap("logo_ico.ico")
    prompt_page.resizable(0,0)
    prompt_page.config(bg='lightgreen')
    
    screen_width = home.winfo_screenwidth()
    screen_height = home.winfo_screenheight()
    window_width = 500  
    window_height = 500  
    x = int((screen_width - window_width) / 2)
    y = int((screen_height - window_height) / 2)
    prompt_page.geometry("+{}+{}".format(x, y))

    prompt_text_widget1 = Text(prompt_page, wrap=WORD, width=44, height=10, font=("Consolas", 10))
    prompt_text_widget1.place(relx=0.5, rely=0.3, anchor=CENTER)
    
    prompt_text_widget2 = Text(prompt_page, wrap=WORD, width=20, height=2,bg='white',fg='black', font=("consolas", 10))
    prompt_text_widget2.place(relx=0.5, rely=0.6, anchor=CENTER)
    prompt_text_widget2.config(state=DISABLED)
    
    generate_button = Button(prompt_page, text="Analyze", font=("consolas", 12),bg= 'blue', fg='white', command=generate)
    generate_button.place(relx=0.3, rely=0.8, anchor=CENTER)
    
    def clear_text():
        prompt_text_widget1.delete("1.0", END)
        prompt_text_widget2.config(state=NORMAL)
        prompt_text_widget2.delete("1.0", END)
        prompt_text_widget2.config(state=DISABLED)
    
    c_button = Button(prompt_page, text="Delete Promt", font=("consolas", 12), bg= 'red', fg='white', command=clear_text)
    c_button.place(relx=0.7, rely=0.8, anchor=CENTER)


    prompt_page.mainloop()






    
    
    
    
# View record:
def history():
    global home2
        
    home2=tk.Tk()
    home2.geometry('500x500')
    home2.iconbitmap("logo_ico.ico")
    home2.title('Feelin - History')
    home2.resizable(0,0)
    home2.config(bg='lightgreen')
    
    screen_width = home2.winfo_screenwidth()
    screen_height = home2.winfo_screenheight()
    window_width = 500  
    window_height = 500  
    x = int((screen_width - window_width) / 2)
    y = int((screen_height - window_height) / 2)
    home2.geometry("+{}+{}".format(x, y))

    df = pd.read_excel(filename)
    
    tree = ttk.Treeview(home2, column=('#c1', '#c2', '#c3'), show='headings', height=15)

    tree.column('#1', width=150, minwidth=150, anchor=tk.CENTER)
    tree.column('#2', width=150, minwidth=150, anchor=tk.CENTER)
    tree.column('#3', width=150, minwidth=150, anchor=tk.CENTER)
    tree.heading('#1', text='Positive')
    tree.heading('#2', text='Neutral')
    tree.heading('#3', text='Negative')
    tree.pack()
    
    for index, row in df.iterrows():
        tree.insert('', 'end', values=tuple('' if pd.isna(x) else x for x in row))

    tree.place(relx=0.5,rely=0.45,anchor=CENTER)
    
    but4 = Button(home2, text='Reset', font=('conslas', 15), command=resetpoints, height=1, width=12, bg='red',
           fg='white', activebackground='Skyblue', activeforeground='thistle1')
    but4.place(relx=0.5,rely=0.9,anchor=CENTER)
    
    home2.mainloop()






# Home Page:
home=tk.Tk()
home.geometry('500x500')
home.iconbitmap("logo_ico.ico")
home.title('Feelin')
home.protocol('WM_DELETE_WINDOW',homeclose)
home.resizable(0,0)
home.config(bg='lightgreen')

screen_width = home.winfo_screenwidth()
screen_height = home.winfo_screenheight()
window_width = 500  
window_height = 500  
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
home.geometry("+{}+{}".format(x, y))

image_home = Image.open("Front.png")
photo_image = ImageTk.PhotoImage(image_home)
label_home = tk.Label(home, image=photo_image)
label_home.config(border='0')
label_home.pack()
label_home.place(relx=0.5,rely=0.2,anchor=CENTER)

but1 = Button(home,text='Analysis',font=('Consolas',20),command=analysis_window,height=1,width=16,bg='green',
    fg='white',activebackground='white',activeforeground='black')
but1.place(relx=0.5,rely=0.5,anchor=CENTER)
but2 = Button(home,text='History',font=('Consolas',20),command=history,height=1,width=16,bg='green',
    fg='white',activebackground='white',activeforeground='black')
but2.place(relx=0.5,rely=0.7,anchor=CENTER)

home.mainloop()







